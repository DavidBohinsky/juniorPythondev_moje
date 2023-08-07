from openpyxl import load_workbook
import json
workbook =load_workbook(filename="reviews-sample.xlsx")

print(workbook.sheetnames)

sheet = workbook.active
#print(sheet["F10"].value)
#print(sheet["A1:C4"])
#print(sheet["A:B"])

#for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):    #vypisanie riadka od 1-2 a stplcov 1-3
#    print(row)

#for column in sheet.iter_cols(min_row=1, max_row=2, min_col=1, max_col=3):
#    print(column)

produkty = {}

for value in sheet.iter_rows(min_row=2, min_col=4, max_col=7, values_only=True):   # vypisujem excel / 4 stlpce
    productId = riadok[0]
    produkt = {
        "parent" : riadok[1],
        "title" : riadok[2],
        "category" : riadok[3],
    }
    produkty[productId] = produkt

print(json.dumps(produkty))
